import random
import string

import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from scipy import stats

listt_m=[] #-------უნიკალური მოდელები 70
C_num=1
for i in range(1,71):

    result = ''.join((random.choice(string.ascii_uppercase) for x in range(5,10)))
    listt_m.append(result)
    for l in listt_m:
        a=listt_m.count(l)
        if a>C_num:
            listt_m.remove(l)
# print(listt_m)


multi_mod=[] # აქ დუბლიკატი მოდელებიც მაქვს ცხრილში რომ ერთნაირი სიგრძის ქოლუმნებ მქონდეს
for i in range(30):
    for m in listt_m:
        # print(m)
        multi_mod.append(m)
# print(len(multi_mod))


listt_sh=[] #-----------შტრიხკოდები ზოგადად უნიკალური
C_num=1
for i in range(2100):

    # result_str = ''.join(random.sample(string.ascii_uppercase, 8)) # აქ სამპლი არ მუშაობს
    result =''.join(random.choice(string.digits) for o in range(11))
    listt_sh.append(result)
    for l in listt_sh:
        a=listt_sh.count(l)
        if a>C_num:
            listt_sh.remove(l)
# print(len(listt_sh))


rand_price=[]
for i in range(2100): #აქ შტრიხკოდების მიხედვით ვითვლი
    n = random.randint(500, 5000)
    rand_price.append(n)
# print(rand_price)

sold_productn=dict()     # უნიკალური მოდელების მიხედვით , ფასები 30 დღის , ეს დიქშნარია სხვანარიად გამოაქვს
# for i in listt_m:
for i in listt_sh:
    for d in range(30):
        n = random.randint(0, 50)
        # print(d)
        sold_productn.setdefault(i, []).append(n)
# print(sold_productn)

#------------------------------აქ უკვე ფრეიმი მინდა
data={'model':multi_mod,
      'shtrixk' :listt_sh,
      'price' : rand_price,
      'sold_m': sold_productn.values()
      }
df=pd.DataFrame(data)
# print(df)

k={"sold_sum": sold_productn.values()} # ეს მინდა დავაჯამო,ანუ ყველა 30 დღის ჯამი მინდა ბოლოს  და შევიტანო მაგიტომ ჩავაინსერტე ბოლოს
add=pd.DataFrame(k)
add["sold_sum"] = add["sold_sum"].apply(lambda x: sum(x) if isinstance(x, (list, tuple)) else x)
 ## აჯამებს # ეს ბოლომდე არ მესმის ინტერნეტი გამოვიყენე მაგრამ შედეგს სწორს იძლევა,
  # ერთ ქოლუმშია ლისტად 30 დღის ინფო და ლისტებში უნდა აჯამოს სათითაოდ  და რთული იყო


df.insert(4, "sold_sum",add["sold_sum"].apply(lambda x: sum(x) if isinstance(x, (list, tuple)) else x) , True)
# print(df)


# df.to_excel("cond.xlsx", sheet_name="sheetOne",index=False)


#------------------

from openpyxl import load_workbook
from openpyxl.workbook import Workbook
# import openpyxl as pxl
# excel_book = pxl.load_workbook('cond.xlsx')
# with pd.ExcelWriter('cond.xlsx', engine='openpyxl') as writer:
#     writer.book = excel_book
#     writer.sheets = {worksheet.title: worksheet for worksheet in excel_book.worksheets}
#
#     mean_p=round(df.price.mean(), 2) #აქ ჯერ საშუალო დავითვალე
#     print(mean_p)
#     new_list=[]
#     for i in range(len(rand_price)):      # აქ ვადარებ და ახალ ლისტს ვქმნი
#         if rand_price[i]>mean_p:
#             new_list.append(rand_price[i])
#     # print(new_list)
#
#
#     data2={'price_mean':new_list}
#     df2=pd.DataFrame(data2)
#     # print(df2)
#     df2.to_excel(writer, 'sheetB', index=False)
#     writer.save()
# df2.to_excel("cond.xlsx", sheet_name="Sheet2") # არ მუშაობს ეს . არ მიაქვს იქ


#------------------------------------------------- სტანდარტული გადახრა,საშუალო
from openpyxl import load_workbook
from openpyxl.workbook import Workbook
# import openpyxl as pxl
# excel_book = pxl.load_workbook('cond.xlsx')
# with pd.ExcelWriter('cond.xlsx', engine='openpyxl') as writer:
#     writer.book = excel_book
#     writer.sheets = {worksheet.title: worksheet for worksheet in excel_book.worksheets}
#
#     from statistics import mean
#     avg_list=[]                            #საშუალოს ითვლის
#     for i in sold_productn.values():
#         avg=round(mean(i),2)
#         avg_list.append(avg)
#     # print(avg_list)
#
#
#     from statistics import pstdev
#     std_list=[]                           #სტანდარტული გადახრა
#     for i in sold_productn.values():
#         std=round(pstdev(i),2)
#         std_list.append(std)
#     # print(std_list)
#
#     data3 = {'mean': avg_list, 'str': std_list}
#     df3 = pd.DataFrame(data3)
#     df3.to_excel(writer, 'sheetG', index=False)
#     writer.save()

#----------------------------------------პერცენტილი
p=df.loc[:,"sold_sum"]
a=np.array(p)
pp= np.percentile(a,75)
# print(pp)

#--------------------------სკატერი
# d=df.loc[:,"sold_sum"]
# plt.scatter(df.index,d, color= "red", s=20, marker='o')
# plt.xlabel("sold_mod")
# plt.ylabel("kondincioneri_mod")
# plt.show()

#-------------------------------------------რეგრესიული
import matplotlib.pyplot as plt
from scipy import stats
# x=df.sold_sum
# y=df.price

# slope, intercept, r, p, std_err = stats.linregress(x, y)
# def funcc_pd(x):
#   return slope * x + intercept
#
# Build_model = list(map(funcc_pd, x))
#
# plt.scatter(x,y, color = "red", s=20, marker='o')
# plt.plot(x, Build_model)
# plt.xlabel("sold_number")
# plt.ylabel("price_of_models")
# plt.show()

#-----------------------------ერთისთვის არის ეს ანუ ეს ერთი მოდელის 30 კონდინციონერისთვის
# mod_1=random.choice(listt_m) # რენდომად ამოიღებს მოდელს
# # print(mod_1)

singl_mod=df[df['model'] == mod_1] #გაიფილტრება ფრეიმი
print(singl_mod)


x=singl_mod.sold_sum
y=singl_mod.price

slope, intercept, r, p, std_err = stats.linregress(x, y)   #r კორელაციაა
def funcc_pd(x):
  return slope * x + intercept

Build_model = list(map(funcc_pd, x))

plt.scatter(x,y, color = "red", s=20, marker='o')
plt.plot(x, Build_model)
plt.xlabel("sold_number")
plt.ylabel("price_of_models")
plt.show()
